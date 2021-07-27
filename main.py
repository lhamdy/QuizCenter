import xlrd
from openpyxl import load_workbook
import time
from random import randint
from datetime import date



workbook = xlrd.open_workbook('QuizCenter.xlsx')
workbook2 = load_workbook('QuizCenter.xlsx')


current_user = { #logs the username and full name of the current user
    'username': '',
    'first name': '',
    'last name': ''
}

#if a new account is made while the code is run, their details are logged here so the user can login after they create the acount
#added this because the excel sheet would not update with the new users details/allow them to login using the normal login process,
#but when a new instance of the code is run, the user will be able to login normally
new_account = {
    'username': '',
    'password': '',
    'first name': '',
    'last name': ''
}

#checks if the user has been registered in the system based off their username
def registered(username):
    return username in workbook.sheet_by_name('user_info').col_values(2)

#registers the new user in the system
def register():
    print('\n\nRegistration page\n')
    first = input('first name: ')
    last = input('last name: ')
    username = input('username: ')
    while username in workbook.sheet_by_name('user_info').col_values(2): #ensures that the user selects a unique username
        print('That username is already taken. Please enter a new username.')
        username = input('username: ')
    password = input('password: ')
    age = input('age: ')
    faculty = input('workplace: ')
    user_info = workbook2['user_info'] #sheet on excel where user info is stored
    info = [first, last, username, password, age, faculty]
    user_info.append(info) #appends new user info to the excel sheet
    workbook2.save(filename='QuizCenter.xlsx')
    print('User registered successfully!\n\n')

    new_account.update({'first name': first}) #updates the new_account with the new users info so they can login immediately
    new_account.update({'last name': last})
    new_account.update({'username': username})
    new_account.update({'password': password})
    time.sleep(2)
    login() #redirects the new user to login



def login():
    print('Welcome to the Quiz Center!\n')
    user = input('username: ')
    password = input('password: ')
    user_info = workbook.sheet_by_name('user_info')
    if user == new_account['username']: #if the username belongs to new user
        if password == new_account['password']: #user login is successful, welcomes them to quiz center
            time.sleep(2)
            print('\nLogin successful!')
            current_user.update({'username': user})
            current_user.update({'first name': new_account['first name']})
            current_user.update({'last name': new_account['last name']})
            print('Welcome, ' + current_user['first name'] + ' ' + current_user['last name'] + '!\n')
            return True

        else: #if the user's password is wrong, prompts them to login again
            print("Wrong Password!\n")
            login()
    elif not registered(user): #if the user isn't in the system, prompts them to create a new account
        regist = input(f'An account with the username "{user}" does not exist. Would you like to register for an account? (y/n) ')
        if regist[0] == 'y':
            register()
        else: #if the user doesn't want to create an account, it exits
            return False
    elif password != user_info.row_values(user_info.col_values(2).index(user))[3]: #if the users password is wrong, prompts them to login again
        print("Wrong Password!\n")
        login()
    else: #user login is success, welcomes them to quiz center
        time.sleep(2)
        print('\nLogin successful!')
        user_row = user_info.row_values(user_info.col_values(2).index(user))
        current_user.update({'username': user})
        current_user.update({'first name': user_row[0]})
        current_user.update({'last name': user_row[1]})
        print('Welcome, ' + current_user['first name'] + ' ' + current_user['last name'] + '!\n')
    return True

def homepage():
    time.sleep(2)
    user_results = workbook2['user_results'] #opens user results sheet
    taken_quizzes = False #if user has taken quizzes or not
    print('\nPrevious quiz results:')
    for row in user_results.iter_rows(values_only=True): #checks if user has taken quizzes. if they have, prints results of previous quiz and date taken
        if row[0] == current_user['username']:
            print(f'Quiz name: {row[1]}')
            print(f'Score: {row[2]}')
            print(f'Date taken: {row[3]}\n')
            taken_quizzes = True
    if not taken_quizzes: #if the user hasn't taken quizzes
        print('You have not taken any quizzes yet. Begin taking quizzes and you will see your results on this page\n')

    time.sleep(2)


def take_quiz():
    print('\nAvailable quizzes:\n')
    available_quizzes = workbook.sheet_by_name('quizzes_files')
    i = 1
    while i < available_quizzes.nrows: #presents user with all available quizzes
        row = available_quizzes.row_values(i)
        print(f'{str(int(row[0]))}. {row[1]}')
        i += 1
    take = ' '
    selected_quiz = '' #quiz the user selected
    quiz_name = '' #name of the sheet where the question bank for the quiz is
    quiz_row = None
    while take[0] != 'y': #prompts the user to enter the quiz they would like to take
        quiz = int(input('\nWhich quiz would you like to take? (Please input the number of the quiz) '))
        quiz_row = available_quizzes.row_values(quiz) #prints quiz description
        print(f'\n{quiz_row[4]}')
        take = input('\nWould you like to take this quiz? (y/n) ') #confirms user's selection, if y proceeds to take quiz, if n prompts them to select a different quiz
        if take[0] == 'y':
            selected_quiz = quiz_row[1]
            quiz_name = quiz_row[6]

    print(f'\nQuiz selected: {selected_quiz}\nLoading quiz...\n\n') #beginning of quiz
    time.sleep(5) #loads quiz

    quiz_sheet = workbook.sheet_by_name(quiz_name) #sheet with questions for quiz
    print('''Instructions: This quiz consists of 10 true/false questions. The questions will appear one at a time, and you must answer each question before
continuing to the next. To input your answers, simply enter 'TRUE' or 'FALSE'. You may now begin the quiz. Good luck!\n''')

    old_results = workbook2['results']
    idx = workbook2.sheetnames.index('results')
    workbook2.remove(old_results) #clears results sheet of previous data
    workbook2.create_sheet('results', idx) #creates a new results sheet for new quiz
    results = workbook2['results']
    results.append(['Question', 'Answer', 'User Answer'])

    rands = []
    for j in range(10): #displays questions to the user in random order
        rand = randint(1, 10)
        while rand in rands: #makes sure question hasn't been asked before
            rand = randint(1, 10)
        rands.append(rand)
        question = quiz_sheet.row_values(rand)
        answer = input(f'{j+1}. {question[1]}: ').upper() #takes user answer
        results.append([question[1], question[2], answer]) #adds user answer to results sheet

    workbook2.save(filename='QuizCenter.xlsx')

    num_correct = 0 #number of questions the user answered correctly
    for row in results.iter_rows(values_only=True): #tallys number of correct answers
        if row[0] == 'Question':
            continue
        if row[1] == row[2]:
            num_correct += 1

    score = (num_correct/10) * 100 #calculates score (percentage)

    user_results = workbook2['user_results']  #logs user results in results sheet

    today = date.today()
    date_taken = today.strftime("%m/%d/%y")

    user_score = [current_user['username'], selected_quiz, f'{score}%', f'{date_taken}']
    user_results.append(user_score)
    workbook2.save(filename='QuizCenter.xlsx')

    time.sleep(2)
    print('\nQuiz complete!')

    time.sleep(2)
    print(f'Score: {score}%')
    if(score > quiz_row[5]):
        print('You passed this quiz! Great job :)\n')
    else:
        print('You failed this quiz :( Better luck next time!\n')

    time.sleep(2)

    review = input('Would you like to review the answers to your quiz? (y/n) ') #asks user if they want to review their quiz

    if review[0] == 'y': #if the user does want to review their results, presents the results to them
        time.sleep(2)
        i = 1
        for row in results.iter_rows(values_only=True):
            if row[0] == 'Question':
                continue
            print(f'\nQuestion {i}: {row[0]}')
            print(f'Correct answer: {row[1]}')
            print(f'Your answer: {row[2]}\n')
            i += 1

q = ''
logged_in = login() #begins login process
while q != 'q': #the user can take quizzes and review scores for previous quizzes

    if logged_in: #user can only access homepage and quizzes if logged in
        homepage()
        take_quiz()
        q = input('Would you like to take more quizzes? (if yes, enter \'y\' to be redirected to the home page. if you would like to logout, enter \'q\') ') #asks user if they would like to continue or log out after each quiz


